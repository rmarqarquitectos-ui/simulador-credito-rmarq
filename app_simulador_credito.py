import streamlit as st
import pandas as pd
from datetime import datetime, date
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

VERSION = "v3.2 UI FIX FINAL RMARQ"

st.set_page_config(
    page_title="Simulador de Crédito",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --------------------------------------------------
# ESTILOS VISUALES FINALES
# --------------------------------------------------
st.markdown("""
<style>
    :root {
        --bg: #f7f9fc;
        --sidebar-bg: #eef3f8;
        --card-bg: #ffffff;
        --text: #1f2937;
        --muted: #6b7280;
        --border: #d9e2ec;
        --primary-soft: #8da2b8;
        --primary-soft-hover: #7f95ac;
        --soft-blue: #eaf3ff;
        --soft-yellow: #fff8db;
        --table-head: #edf3f8;
        --table-row: #ffffff;
    }

    .stApp {
        background: var(--bg);
        color: var(--text);
    }

    header[data-testid="stHeader"] {
        background: transparent !important;
        height: 0 !important;
    }

    .stAppToolbar,
    div[data-testid="stDecoration"] {
        display: none !important;
    }

    section[data-testid="stSidebar"] {
        width: 355px !important;
        min-width: 355px !important;
        background: var(--sidebar-bg) !important;
        border-right: 1px solid var(--border);
    }

    section[data-testid="stSidebar"] > div {
        padding-top: 0.4rem !important;
        padding-left: 0.85rem !important;
        padding-right: 0.85rem !important;
        padding-bottom: 0.45rem !important;
    }

    .block-container {
        max-width: 1320px;
        padding-top: 0.65rem !important;
        padding-bottom: 1rem !important;
    }

    h1 {
        color: var(--text) !important;
        font-size: 1.95rem !important;
        font-weight: 800 !important;
        margin-bottom: 0.8rem !important;
        line-height: 1.1 !important;
    }

    h2, h3 {
        color: var(--text) !important;
        font-weight: 780 !important;
        margin-top: 0.55rem !important;
        margin-bottom: 0.45rem !important;
        line-height: 1.1 !important;
    }

    p, label, span, div, li {
        color: var(--text);
    }

    .sidebar-title {
        font-size: 1.25rem;
        font-weight: 800;
        color: var(--text);
        margin-bottom: 0.32rem;
        line-height: 1.05;
    }

    .sidebar-section {
        font-size: 0.94rem;
        font-weight: 760;
        color: var(--text);
        margin-top: 0.32rem;
        margin-bottom: 0.08rem;
        line-height: 1.05;
    }

    .sidebar-note {
        font-size: 0.73rem;
        color: var(--muted);
        margin-top: 0rem;
        margin-bottom: 0.12rem;
        line-height: 1.15;
    }

    .stSidebar label {
        font-size: 0.80rem !important;
        margin-bottom: 0.04rem !important;
    }

    .stSidebar [data-testid="stVerticalBlock"] > div {
        margin-bottom: 0.02rem !important;
    }

    .stSidebar .stTextInput,
    .stSidebar .stCheckbox,
    .stSidebar .stRadio {
        margin-bottom: 0.08rem !important;
    }

    .stTextInput input {
        background: #ffffff !important;
        color: #111827 !important;
        -webkit-text-fill-color: #111827 !important;
        border: 1px solid var(--border) !important;
        border-radius: 10px !important;
        font-size: 0.84rem !important;
        min-height: 30px !important;
        height: 30px !important;
        padding-top: 0.08rem !important;
        padding-bottom: 0.08rem !important;
        box-shadow: none !important;
    }

    .stTextInput > div,
    .stTextInput > div > div {
        background: transparent !important;
        box-shadow: none !important;
    }

    .stTextInput input:focus {
        box-shadow: none !important;
        outline: none !important;
        border: 1px solid #b8c6d6 !important;
    }

    .stTextInput input::placeholder {
        color: #8a94a6 !important;
        opacity: 1 !important;
    }

    .stRadio [role="radiogroup"] {
        display: flex !important;
        gap: 0.45rem !important;
        margin-top: 0.05rem !important;
        margin-bottom: 0.05rem !important;
    }

    .stRadio label {
        background: #ffffff !important;
        border: 1px solid var(--border) !important;
        border-radius: 10px !important;
        padding: 0.28rem 0.72rem !important;
        min-height: 30px !important;
        display: flex !important;
        align-items: center !important;
        box-shadow: none !important;
    }

    .stRadio label div {
        color: #111827 !important;
        font-size: 0.84rem !important;
    }

    .stCheckbox label {
        font-size: 0.82rem !important;
    }

    div[data-testid="metric-container"] {
        background: var(--card-bg);
        border: 1px solid var(--border);
        border-radius: 15px;
        padding: 0.85rem 0.95rem 0.75rem 0.95rem;
        box-shadow: 0 2px 10px rgba(15, 23, 42, 0.04);
    }

    div[data-testid="stMetricLabel"] {
        color: var(--muted) !important;
        font-size: 0.82rem !important;
        font-weight: 600 !important;
    }

    div[data-testid="stMetricValue"] {
        color: var(--text) !important;
        font-size: 1.55rem !important;
        font-weight: 800 !important;
        line-height: 1.05 !important;
    }

    [data-testid="stAlert"] {
        border-radius: 12px !important;
        border: 1px solid var(--border) !important;
    }

    .stDownloadButton button {
        background: var(--primary-soft) !important;
        color: #ffffff !important;
        border: none !important;
        border-radius: 10px !important;
        padding: 0.5rem 0.9rem !important;
        font-weight: 700 !important;
        font-size: 0.88rem !important;
        box-shadow: none !important;
    }

    .stDownloadButton button:hover {
        background: var(--primary-soft-hover) !important;
        color: #ffffff !important;
    }

    .table-wrap {
        background: #ffffff;
        border: 1px solid var(--border);
        border-radius: 16px;
        overflow: hidden;
        box-shadow: 0 2px 10px rgba(15, 23, 42, 0.04);
    }

    table.credit-table {
        width: 100%;
        border-collapse: collapse;
        font-size: 0.9rem;
    }

    .credit-table thead th {
        background: var(--table-head);
        color: var(--text);
        text-align: left;
        padding: 11px 13px;
        border-bottom: 1px solid var(--border);
        font-weight: 760;
        white-space: nowrap;
    }

    .credit-table tbody td {
        background: var(--table-row);
        color: var(--text);
        padding: 10px 13px;
        border-bottom: 1px solid #edf2f7;
        white-space: nowrap;
    }

    .credit-table tbody tr:hover td {
        background: #fafcff;
    }

    .credit-table tbody tr:last-child td {
        border-bottom: none;
    }

    .small-caption {
        color: var(--muted);
        font-size: 0.8rem;
        margin-top: 0.3rem;
        line-height: 1.25;
    }

/* Fix definitivo: eliminar rectángulo oscuro en inputs */
.stTextInput, .stTextInput * {
    background: transparent !important;
    box-shadow: none !important;
}

.stTextInput div[data-baseweb="input"],
.stTextInput div[data-baseweb="base-input"]{
    background: transparent !important;
    box-shadow: none !important;
    border: none !important;
}

.stTextInput input {
    background: #ffffff !important;
    color: #111827 !important;
    border: 1px solid #d9e2ec !important;
    box-shadow: none !important;
}

.stSidebar .stTextInput {
    background: transparent !important;
}

/* Eliminar cualquier sombra inferior tipo rectángulo */
.stTextInput > div {
    box-shadow: none !important;
    border: none !important;
}

</style>
""", unsafe_allow_html=True)

# --------------------------------------------------
# FUNCIONES
# --------------------------------------------------
def formato_pesos(valor):
    return f"${float(valor):,.0f}".replace(",", ".")

def formato_uf(valor):
    return f"{float(valor):,.3f}".replace(",", ".")

def parsear_numero(texto, decimales_permitidos=True):
    texto = (texto or "").strip()
    if texto == "":
        return 0.0

    texto = texto.replace(" ", "")

    # Caso 1: tiene coma y punto, asumimos punto miles y coma decimal
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    # Caso 2: solo coma, asumimos coma decimal
    elif "," in texto:
        texto = texto.replace(",", ".")
    # Caso 3: solo punto, lo dejamos como decimal normal
    else:
        texto = texto

    try:
        valor = float(texto)
        if valor < 0:
            return "ERROR"
        if not decimales_permitidos:
            return float(int(round(valor)))
        return valor
    except ValueError:
        return "ERROR"

def parsear_entero(texto):
    valor = parsear_numero(texto, decimales_permitidos=False)
    if valor == "ERROR":
        return "ERROR"
    return int(valor)

def parsear_fecha_texto(fecha_texto):
    fecha_texto = (fecha_texto or "").strip()
    if fecha_texto == "":
        return None
    try:
        return datetime.strptime(fecha_texto, "%d/%m/%Y").date()
    except ValueError:
        return "ERROR"

def construir_simulacion(
    tipo_credito,
    monto_base,
    tasa_anual_pct,
    fecha_primera_cuota,
    ite,
    gastos_notariales,
    incluir_gastos,
    cuotas_solo_interes,
    amortizacion_cuota_4,
    valor_uf
):
    tasa = float(tasa_anual_pct) / 100.0

    monto_total = float(monto_base)
    if incluir_gastos:
        monto_total += float(ite) + float(gastos_notariales)

    saldo = monto_total
    filas = []

    for cuota_num in range(1, 6):
        saldo_inicial = float(saldo)
        interes = saldo_inicial * tasa

        if cuota_num <= int(cuotas_solo_interes):
            amortizacion = 0.0
        elif cuota_num == 4:
            amortizacion = min(float(amortizacion_cuota_4), saldo_inicial)
        else:
            amortizacion = saldo_inicial

        cuota_total = interes + amortizacion
        saldo_final = saldo_inicial - amortizacion

        if fecha_primera_cuota is None:
            fecha_txt = ""
        else:
            fecha_real = date(
                fecha_primera_cuota.year + (cuota_num - 1),
                fecha_primera_cuota.month,
                fecha_primera_cuota.day
            )
            fecha_txt = fecha_real.strftime("%d-%m-%Y")

        fila = {
            "Cuota": cuota_num,
            "Fecha": fecha_txt,
            "Saldo inicial": saldo_inicial,
            "Interés": interes,
            "Amortización": amortizacion,
            "Cuota total": cuota_total,
            "Saldo final": saldo_final
        }

        if tipo_credito == "UF" and valor_uf > 0:
            fila["Saldo inicial ($)"] = saldo_inicial * valor_uf
            fila["Interés ($)"] = interes * valor_uf
            fila["Amortización ($)"] = amortizacion * valor_uf
            fila["Cuota total ($)"] = cuota_total * valor_uf
            fila["Saldo final ($)"] = saldo_final * valor_uf

        filas.append(fila)
        saldo = saldo_final

    df = pd.DataFrame(filas)
    total_pagado = float(df["Cuota total"].sum())
    interes_total = total_pagado - monto_total
    cuoton_capital = float(df.iloc[-1]["Amortización"])

    return df, monto_total, total_pagado, interes_total, cuoton_capital

def dataframe_a_html(df, tipo_credito):
    columnas = list(df.columns)
    html = '<div class="table-wrap"><table class="credit-table">'
    html += "<thead><tr>"
    for col in columnas:
        html += f"<th>{col}</th>"
    html += "</tr></thead><tbody>"

    for _, row in df.iterrows():
        html += "<tr>"
        for col in columnas:
            valor = row[col]
            if isinstance(valor, (int, float)):
                if tipo_credito == "UF" and "($)" not in col and col != "Cuota":
                    texto = formato_uf(valor)
                elif col == "Cuota":
                    texto = str(int(valor))
                else:
                    texto = formato_pesos(valor)
            else:
                texto = str(valor)
            html += f"<td>{texto}</td>"
        html += "</tr>"
    html += "</tbody></table></div>"
    return html

def generar_excel(parametros, resumen, df):
    salida = BytesIO()
    with pd.ExcelWriter(salida, engine="openpyxl") as writer:
        sheet_name = "Simulacion"

        fila_parametros = pd.DataFrame(parametros, columns=["Parámetro", "Valor"])
        fila_resumen = pd.DataFrame(resumen, columns=["Resumen", "Valor"])

        fila_parametros.to_excel(writer, index=False, sheet_name=sheet_name, startrow=0)
        fila_resumen.to_excel(writer, index=False, sheet_name=sheet_name, startrow=len(fila_parametros) + 3)
        df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=len(fila_parametros) + len(fila_resumen) + 7)

        ws = writer.sheets[sheet_name]

        # Estilos
        fill_header = PatternFill("solid", fgColor="EAF3FF")
        fill_subheader = PatternFill("solid", fgColor="EDF3F8")
        bold = Font(bold=True)
        border = Border(
            left=Side(style="thin", color="D9E2EC"),
            right=Side(style="thin", color="D9E2EC"),
            top=Side(style="thin", color="D9E2EC"),
            bottom=Side(style="thin", color="D9E2EC")
        )

        # Anchos
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 18
        ws.column_dimensions["H"].width = 18
        ws.column_dimensions["I"].width = 18
        ws.column_dimensions["J"].width = 18
        ws.column_dimensions["K"].width = 18
        ws.column_dimensions["L"].width = 18

        # Encabezados secciones
        for cell in ws[1]:
            cell.fill = fill_header
            cell.font = bold
            cell.border = border

        resumen_inicio = len(fila_parametros) + 4
        for cell in ws[resumen_inicio]:
            cell.fill = fill_header
            cell.font = bold
            cell.border = border

        tabla_inicio = len(fila_parametros) + len(fila_resumen) + 8
        for cell in ws[tabla_inicio]:
            cell.fill = fill_subheader
            cell.font = bold
            cell.border = border

        # Bordes al contenido
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=min(ws.max_column, 12)):
            for cell in row:
                if cell.value is not None:
                    cell.border = border
                    cell.alignment = Alignment(vertical="center")

    salida.seek(0)
    return salida.getvalue()

# --------------------------------------------------
# SIDEBAR
# --------------------------------------------------
st.sidebar.markdown('<div class="sidebar-title">Simulador</div>', unsafe_allow_html=True)

st.sidebar.markdown('<div class="sidebar-section">Tipo de crédito</div>', unsafe_allow_html=True)
tipo_credito = st.sidebar.radio(
    "Unidad del crédito",
    ["Pesos", "UF"],
    horizontal=True
)

if tipo_credito == "UF":
    valor_uf_txt = st.sidebar.text_input("Valor UF ($)", value="", placeholder="0")
    valor_uf = parsear_numero(valor_uf_txt, decimales_permitidos=True)
    if valor_uf == "ERROR":
        st.sidebar.error("Valor UF inválido.")
        valor_uf = 0.0
    st.sidebar.markdown(
        '<div class="sidebar-note">Todos los valores del crédito deben ingresarse en UF.</div>',
        unsafe_allow_html=True
    )
else:
    valor_uf = 0.0
    st.sidebar.markdown(
        '<div class="sidebar-note">Todos los valores del crédito deben ingresarse en pesos.</div>',
        unsafe_allow_html=True
    )

st.sidebar.markdown('<div class="sidebar-section">Parámetros</div>', unsafe_allow_html=True)

monto_base_txt = st.sidebar.text_input(
    f"Monto base ({'UF' if tipo_credito == 'UF' else '$'})",
    value="",
    placeholder="0"
)
monto_base = parsear_numero(monto_base_txt, decimales_permitidos=(tipo_credito == "UF"))
if monto_base == "ERROR":
    st.sidebar.error("Monto base inválido.")
    monto_base = 0.0

tasa_txt = st.sidebar.text_input(
    "Tasa anual (%)",
    value="",
    placeholder="0"
)
tasa_anual_pct = parsear_numero(tasa_txt, decimales_permitidos=True)
if tasa_anual_pct == "ERROR":
    st.sidebar.error("Tasa inválida.")
    tasa_anual_pct = 0.0

fecha_texto = st.sidebar.text_input(
    "Fecha 1ª cuota",
    value="",
    placeholder="Día/Mes/Año"
)
fecha_primera_cuota = parsear_fecha_texto(fecha_texto)
if fecha_primera_cuota == "ERROR":
    st.sidebar.error("La fecha debe escribirse como Día/Mes/Año, por ejemplo 30/08/2027.")
    fecha_primera_cuota = None

st.sidebar.markdown('<div class="sidebar-section">Gastos</div>', unsafe_allow_html=True)

ite_txt = st.sidebar.text_input(
    f"ITE ({'UF' if tipo_credito == 'UF' else '$'})",
    value="",
    placeholder="0"
)
ite = parsear_numero(ite_txt, decimales_permitidos=(tipo_credito == "UF"))
if ite == "ERROR":
    st.sidebar.error("ITE inválido.")
    ite = 0.0

gastos_notariales_txt = st.sidebar.text_input(
    f"Gastos notariales ({'UF' if tipo_credito == 'UF' else '$'})",
    value="",
    placeholder="0"
)
gastos_notariales = parsear_numero(gastos_notariales_txt, decimales_permitidos=(tipo_credito == "UF"))
if gastos_notariales == "ERROR":
    st.sidebar.error("Gastos notariales inválidos.")
    gastos_notariales = 0.0

incluir_gastos = st.sidebar.checkbox("Incluir gastos en el crédito", value=False)

st.sidebar.markdown('<div class="sidebar-section">Estructura</div>', unsafe_allow_html=True)

cuotas_solo_interes_txt = st.sidebar.text_input(
    "Cuotas solo interés",
    value="",
    placeholder="0"
)
cuotas_solo_interes = parsear_entero(cuotas_solo_interes_txt)
if cuotas_solo_interes == "ERROR":
    st.sidebar.error("Cuotas solo interés inválidas.")
    cuotas_solo_interes = 0
if cuotas_solo_interes > 4:
    st.sidebar.error("El máximo permitido es 4.")
    cuotas_solo_interes = 4

amortizacion_cuota_4_txt = st.sidebar.text_input(
    f"Amortización cuota 4 ({'UF' if tipo_credito == 'UF' else '$'})",
    value="",
    placeholder="0"
)
amortizacion_cuota_4 = parsear_numero(amortizacion_cuota_4_txt, decimales_permitidos=(tipo_credito == "UF"))
if amortizacion_cuota_4 == "ERROR":
    st.sidebar.error("Amortización cuota 4 inválida.")
    amortizacion_cuota_4 = 0.0

# --------------------------------------------------
# CÁLCULO
# --------------------------------------------------
df, monto_total, total_pagado, interes_total, cuoton_capital = construir_simulacion(
    tipo_credito=tipo_credito,
    monto_base=monto_base,
    tasa_anual_pct=tasa_anual_pct,
    fecha_primera_cuota=fecha_primera_cuota,
    ite=ite,
    gastos_notariales=gastos_notariales,
    incluir_gastos=incluir_gastos,
    cuotas_solo_interes=cuotas_solo_interes,
    amortizacion_cuota_4=amortizacion_cuota_4,
    valor_uf=valor_uf
)

# --------------------------------------------------
# CONTENIDO PRINCIPAL
# --------------------------------------------------
st.title("Simulador de Crédito")
st.caption(VERSION)

st.subheader("Resumen")
col1, col2, col3 = st.columns(3)

if tipo_credito == "Pesos":
    col1.metric("Monto financiado ($)", formato_pesos(monto_total))
    col2.metric("Total pagado ($)", formato_pesos(total_pagado))
    col3.metric("Interés total ($)", formato_pesos(interes_total))
else:
    col1.metric("Monto financiado (UF)", formato_uf(monto_total))
    col2.metric("Total pagado (UF)", formato_uf(total_pagado))
    col3.metric("Interés total (UF)", formato_uf(interes_total))
    if valor_uf > 0:
        st.caption(f"Equivalente referencial del monto financiado: {formato_pesos(monto_total * valor_uf)}")

if incluir_gastos:
    st.success("En esta simulación, ITE y gastos notariales sí están incorporados al crédito.")
else:
    st.info("En esta simulación, ITE y gastos notariales los paga el cliente fuera del crédito.")

st.subheader("Cuotón final")

if tipo_credito == "Pesos":
    st.warning(
        f"Capital del cuotón final: {formato_pesos(cuoton_capital)}. "
        f"Este valor corresponde solo al capital. La cuota final incluye además el interés del período."
    )
else:
    texto = f"Capital del cuotón final: {formato_uf(cuoton_capital)} UF."
    if valor_uf > 0:
        texto += f" Equivalente referencial: {formato_pesos(cuoton_capital * valor_uf)}."
    texto += " Este valor corresponde solo al capital. La cuota final incluye además el interés del período."
    st.warning(texto)

st.subheader("Tabla de amortización")
st.markdown(dataframe_a_html(df, tipo_credito), unsafe_allow_html=True)

st.subheader("Descarga")

parametros_excel = [
    ["Tipo de crédito", tipo_credito],
    ["Monto base", formato_uf(monto_base) if tipo_credito == "UF" else formato_pesos(monto_base)],
    ["Tasa anual (%)", f"{tasa_anual_pct:.3f}"],
    ["Fecha 1ª cuota", fecha_texto if fecha_texto.strip() else ""],
    ["ITE", formato_uf(ite) if tipo_credito == "UF" else formato_pesos(ite)],
    ["Gastos notariales", formato_uf(gastos_notariales) if tipo_credito == "UF" else formato_pesos(gastos_notariales)],
    ["Incluir gastos en el crédito", "Sí" if incluir_gastos else "No"],
    ["Cuotas solo interés", str(cuotas_solo_interes)],
    ["Amortización cuota 4", formato_uf(amortizacion_cuota_4) if tipo_credito == "UF" else formato_pesos(amortizacion_cuota_4)],
]

if tipo_credito == "UF":
    parametros_excel.insert(1, ["Valor UF", formato_pesos(valor_uf)])

resumen_excel = [
    ["Monto financiado", formato_uf(monto_total) if tipo_credito == "UF" else formato_pesos(monto_total)],
    ["Total pagado", formato_uf(total_pagado) if tipo_credito == "UF" else formato_pesos(total_pagado)],
    ["Interés total", formato_uf(interes_total) if tipo_credito == "UF" else formato_pesos(interes_total)],
    ["Cuotón final capital", formato_uf(cuoton_capital) if tipo_credito == "UF" else formato_pesos(cuoton_capital)],
]

excel_bytes = generar_excel(parametros_excel, resumen_excel, df)

st.download_button(
    label="Descargar tabla en Excel",
    data=excel_bytes,
    file_name="simulacion_credito.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

st.markdown(
    '<div class="small-caption">El Excel incluye parámetros utilizados, resumen de resultados y tabla de amortización.</div>',
    unsafe_allow_html=True
)